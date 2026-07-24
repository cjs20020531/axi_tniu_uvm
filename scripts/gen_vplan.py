#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# =============================================================================
# File        : gen_vplan.py
# Description : Generate the AXI Target NIU (axi_tniu) UVM verification plan
#               workbook (axi_tniu_vplan.xlsx). The plan is derived from the
#               RKNP protocol spec and the axi_tniu detailed-design document,
#               and is scoped to the user's verification configuration:
#                 NBYTEPERWORD=8, HeadPenalty=0, Req-Error ON, Wrap-realign ON,
#                 Early/Bufferable ON, Same-address ordering ON, AXID-map ON.
# Author      : Verification Team
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- shared styles ----------------------------------------------------------
FONT      = "Arial"
H_FILL    = PatternFill("solid", fgColor="1F4E78")   # header dark blue
SUB_FILL  = PatternFill("solid", fgColor="D9E1F2")   # section light blue
CFG_FILL  = PatternFill("solid", fgColor="FFF2CC")   # config yellow
P0_FILL   = PatternFill("solid", fgColor="F8CBAD")   # priority P0
title_font = Font(name=FONT, size=16, bold=True, color="FFFFFF")
hdr_font   = Font(name=FONT, size=10, bold=True, color="FFFFFF")
cell_font  = Font(name=FONT, size=10)
bold_font  = Font(name=FONT, size=10, bold=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
WRAP_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = H_FILL
        cell.alignment = WRAP_C; cell.border = BORDER

def put(ws, r, c, v, font=None, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font or cell_font
    if fill: cell.fill = fill
    cell.alignment = align or WRAP_L
    cell.border = BORDER
    return cell

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# =============================================================================
# Sheet 1 : Cover / 概述
# =============================================================================
ws = wb.active; ws.title = "0-概述Overview"
set_widths(ws, [26, 90])
ws.merge_cells("A1:B1")
put(ws, 1, 1, "AXI Target NIU (axi_tniu)  UVM 验证计划 / Verification Plan",
    title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 34

rows = [
    ("项目 Project",              "RKNoC — AXI Target NIU (RKNP <-> AXI4 协议转换)"),
    ("被测模块 DUT",              "axi_tniu (top), 含 wrap_align/wrap_adjust/req_order/rsp_order/"
                                  "addr_map/rreq_trans/wreq_trans/rsp_trans/watchdog/ely_rsp_detect"),
    ("验证方法学 Methodology",    "UVM 1.2  (SystemVerilog)"),
    ("仿真器 Simulator",          "Synopsys VCS + Verdi (波形/调试)"),
    ("文档依据 References",       "RKNP 详细设计方案书 V1.0；axi_tniu 详细设计方案书"),
    ("", ""),
    ("== 本验证计划的配置 Verification Configuration ==", ""),
    ("NBYTEPERWORD",              "8   (流控层/AXI 数据位宽 = 8*8 = 64 bit)"),
    ("HeadPenalty",               "0   (RKNP 固定；流控层 data 位宽 = head + 1 + 9*NBYTEPERWORD)"),
    ("Req-Error 处理",            "开启 (请求阶段 Status=ERR 的请求需自组织 error response, 不下发 AXI)"),
    ("Wrap 重排 WRAP_ALIGN_MODE", "开启 (=1；非对齐 WRAP 读写地址对齐 + wrap_adjust 响应修正)"),
    ("Early/Bufferable 模式",     "开启 EARLY_RSP_MODE=1 (bufferable 写 early response + ely_rsp_detect)"),
    ("同地址保序 ADDR_BP_TYPE",   "开启 =2 (写后读/写后写/读后写同地址保序; 颗粒度 ADDR_BLOCK_SIZE=64B)"),
    ("AXID 映射",                 "开启 (ORDKEY_WITH=8, AXID_WITH=4, 输入 OrderKey 位宽为 TNIU AXID 的 2 倍)"),
    ("SUP_REQ_NUM",               "8   (最大 outstanding 请求个数 -> TAG_CNT_WITH=$clog2(8)=3)"),
    ("TIMOUT_VALUE",              "1024 (watchdog 超时门限, 可配)"),
    ("", ""),
    ("== 验证目标 Goals ==", ""),
    ("功能覆盖率 Functional Cov", ">= 100% (covergroup 定义见 5-覆盖率计划)"),
    ("代码覆盖率 Code Cov",       ">= 98% line/cond/toggle/fsm/branch (排除项需评审)"),
    ("断言 Assertion",            "接口协议断言(RKNP/AXI4) 全通过, 0 fail"),
    ("回归 Regression",           "全部 directed + random 用例 0 error / 0 fatal"),
    ("", ""),
    ("== Sheet 索引 ==", ""),
    ("1-特性清单",                "DUT 特性分解, 每条特性映射到验证方式与用例ID"),
    ("2-测试用例",                "定向 + 随机用例清单, 优先级/依赖/状态"),
    ("3-验证环境架构",            "UVM 平台组件说明与连接关系"),
    ("4-断言检查点",              "接口协议断言与 scoreboard 检查点"),
    ("5-覆盖率计划",              "功能覆盖 covergroup / coverpoint / cross"),
    ("6-验证进度清单",            "里程碑与逐项 checklist"),
]
r = 3
for k, v in rows:
    if k.startswith("=="):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        put(ws, r, 1, k, bold_font, SUB_FILL, WRAP_L)
    elif k == "" and v == "":
        r += 1; continue
    else:
        put(ws, r, 1, k, bold_font, CFG_FILL if k and not k.startswith("==") and r < 20 else None)
        put(ws, r, 2, v)
    r += 1

# =============================================================================
# Sheet 2 : Feature list / 特性清单
# =============================================================================
ws = wb.create_sheet("1-特性清单FeatureList")
cols = ["Feature ID", "所属模块 Block", "特性 Feature", "描述 Description",
        "配置相关 Config", "优先级", "验证方式 Method", "关联用例 TC-ID"]
set_widths(ws, [11, 14, 24, 46, 16, 8, 16, 18])
ws.merge_cells("A1:H1")
put(ws, 1, 1, "AXI Target NIU 特性分解清单 (Feature List)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(cols, 1): put(ws, 2, i, h)
style_header(ws, 2, len(cols))

feat = [
    # id, block, feature, desc, config, prio, method, tc
    ("F-REQ-01","rreq_trans","读请求解包→AR","RD/RDW request head 编码为 AXI AR 通道信号(araddr/arlen/arsize/arburst/arid...)","基础","P0","定向+随机+SB","TC-RD-*"),
    ("F-REQ-02","wreq_trans","写请求解包→AW/W","WR/WRW request head→AW, body→W(wdata/wstrb/wlast)","基础","P0","定向+随机+SB","TC-WR-*"),
    ("F-REQ-03","addr_map","地址映射","RKNP subrange/address→AXI 40bit 地址; 生成 len/qos","基础","P0","定向+随机+SB","TC-ADDR-*"),
    ("F-REQ-04","addr_map","突发类型/size 生成","按 Len 与 NBYTEPERWORD=8 生成 awsize/arsize=3, INCR/WRAP burst","NBYTE=8","P0","定向+SB","TC-BURST-*"),
    ("F-RSP-01","rsp_trans","R 通道→读响应","AXI R 通道数据编码为 RKNP read rsp packet(含 LW/Be/Byte)","基础","P0","定向+随机+SB","TC-RD-*"),
    ("F-RSP-02","rsp_trans","B 通道→写响应","AXI B 通道编码为 RKNP write rsp packet","基础","P0","定向+随机+SB","TC-WR-*"),
    ("F-RSP-03","rsp_trans","R/B 轮询合并","R/B 通道在 TNIU 内合并, 需轮询读取, 匹配需带读写类型","基础","P1","随机+SB","TC-MIX-*"),
    ("F-FC-01","top(接口)","流控层 HeadPenalty=0","head 与首拍 body 同拍; data 位宽=head+1+9*8; valid/ready 握手","HP=0","P0","协议断言","TC-FC-*"),
    ("F-FC-02","top(接口)","head/tail 标志","head 标识首拍, tail 标识最后一拍 body flit","基础","P0","协议断言","TC-FC-*"),
    ("F-QOS-01","addr_map/req","QoS/Urgency 编解码","7bit bar-graph Urgency <-> 3bit QoS <-> awqos/arqos","基础","P1","定向+随机+SB","TC-QOS-*"),
    ("F-OUT-01","req/rsp_order","AXI outstanding","多笔请求未回响应即可继续下发, 上限=SUP_REQ_NUM=8","基础","P0","随机+SB","TC-OUT-*"),
    ("F-OOO-01","rsp_order","乱序响应匹配","target 乱序返回响应, 按 {AXID,Opc,TAG_CNT} 索引匹配 head 条目","基础","P0","定向+随机+SB","TC-OOO-*"),
    ("F-ILV-01","rsp_order","读交织处理","同 R 通道多 ID 交织返回, 按交织片段拆分打包; Status=CONT/OK","基础","P0","定向+SB","TC-ILV-*"),
    ("F-ILV-02","rsp_order","交织地址更新","交织中断再续传时重发 rsp head 并更新 Addr","基础","P1","定向+SB","TC-ILV-*"),
    ("F-ILV-03","rsp_trans","LW 生成","burst 最后一拍(rlast=1)所在 cycle 置 LW=1","基础","P1","定向+SB","TC-ILV-*"),
    # ---- enabled optional features ----
    ("F-AXID-01","req_order","AXID 映射(高→低)","OrderKey(8bit)→AXID(4bit) 映射; 多 OrderKey 复用同 AXID","AXID-map ON","P0","定向+随机+SB","TC-AXID-*"),
    ("F-AXID-02","req_order","映射后回填","rsp 阶段用映射后 AXID+TAG_CNT 索引, 回填原始 OrderKey 到 rsp head","AXID-map ON","P0","定向+SB","TC-AXID-*"),
    ("F-AXID-03","req_order","AXID 资源耗尽反压","可用 AXID/TAG 用尽时对上游 rknp_rxreq_ready 反压","AXID-map ON","P1","定向","TC-AXID-BP"),
    ("F-ERR-01","req/rsp_order","请求阶段错误处理","Status=ERR 请求不下发 AXI, 由 rsp_order 自组织 error response","Req-Err ON","P0","定向+SB","TC-ERR-*"),
    ("F-ERR-02","rsp_order","ErrorCode 透传","error rsp 的 ErrorCode 与请求一致(spec_req_buffer)","Req-Err ON","P0","定向+SB","TC-ERR-*"),
    ("F-ERR-03","rsp_order","ERR/OKEY 混合保序","同 AXID 下 err req 与 okey req 混合, follo_err_en/fir_err_en 逻辑","Req-Err ON","P0","定向+随机+SB","TC-ERR-MIX"),
    ("F-ERR-04","rsp_order","错误响应 body 填充","为 err/timeout/early 响应填充 body(LW/Be/Byte)","Req-Err ON","P1","定向+SB","TC-ERR-*"),
    ("F-WRAP-01","wrap_align","非对齐 WRAP 对齐","地址非对齐 WRAP 读写按 RWRAP 边界对齐, 拆分下发","Wrap ON","P0","定向+SB","TC-WRAP-*"),
    ("F-WRAP-02","wrap_adjust","WRAP 响应地址修正","wrap 读/写响应地址按 offset_addr 修正回原始地址","Wrap ON","P0","定向+SB","TC-WRAP-*"),
    ("F-WRAP-03","wrap_adjust","WRAP body 拼接","缓存/拼接首拍数据修正 wrap 读响应 body; 计数上限 RWRAP_CNT_MAX=4","Wrap ON","P1","定向+SB","TC-WRAP-*"),
    ("F-BP-01","req_order","写后读同地址保序","同地址(块 64B)WAR: 读需等前写完成","BP=2 ON","P0","定向+SB","TC-BP-WAR"),
    ("F-BP-02","req_order","写后写同地址保序","同地址 WAW 保序","BP=2 ON","P0","定向+SB","TC-BP-WAW"),
    ("F-BP-03","req_order","读后写同地址保序","同地址 RAW 保序(BP_TYPE=2 特有)","BP=2 ON","P0","定向+SB","TC-BP-RAW"),
    ("F-BP-04","req_order","地址块颗粒检测","按 ADDR_BLOCK_SIZE=64B 块比较, 边界跨块场景","BP=2 ON","P1","定向+随机","TC-BP-BLK"),
    ("F-ELY-01","ely_rsp_detect","bufferable 写 early rsp","bufferable 写请求本地提前回响应, 请求仍下发 target","Early ON","P0","定向+SB","TC-ELY-*"),
    ("F-ELY-02","ely_rsp_detect","real rsp 旁路","bufferable 写的 real response 经 ELY_RSP_TABLE 识别并旁路","Early ON","P0","定向+SB","TC-ELY-*"),
    ("F-ELY-03","ely_rsp_detect","non-bufferable 不 early","non-bufferable 写按常规 real response 处理","Early ON","P0","定向+SB","TC-ELY-NB"),
    ("F-ELY-04","ely_rsp_detect","bufferable 写后读保序","early 模式下写后读需同地址保序(依赖 F-BP-01)","Early ON","P1","定向+SB","TC-ELY-WAR"),
    ("F-ELY-05","ely_rsp_detect","ELY_RSP_TABLE 满","table 满(深度=SUP_REQ_NUM)时行为","Early ON","P2","定向","TC-ELY-FULL"),
    ("F-WD-01","watchdog","超时检测","常规请求超时(TIMOUT_VALUE)返回 timeout error response(ErrorCode=110)","基础","P0","定向+SB","TC-WD-*"),
    ("F-WD-02","watchdog","timer 双倍计数","timer_cnt_max=2*timout_max, req/rsp 计数值换算","基础","P1","定向","TC-WD-CNT"),
    ("F-WD-03","watchdog","超时后 real rsp 丢弃","超时后 target 迟到的 real response 需被丢弃","基础","P1","定向+SB","TC-WD-LATE"),
    ("F-RST-01","top","复位","aresetn 异步复位, 复位中/后接口安全(无非法 valid)","基础","P0","定向+断言","TC-RST-*"),
    ("F-BP2-01","top","反压/背靠背","上游 valid 抖动 + 下游 ready 抖动, 全通道随机反压","基础","P1","随机","TC-STRESS-*"),
]
r = 3
for f in feat:
    prio_fill = P0_FILL if f[5] == "P0" else None
    put(ws, r, 1, f[0], bold_font)
    put(ws, r, 2, f[1]); put(ws, r, 3, f[2]); put(ws, r, 4, f[3])
    put(ws, r, 5, f[4], cell_font, CFG_FILL if "ON" in f[4] else None)
    put(ws, r, 6, f[5], bold_font, prio_fill, WRAP_C)
    put(ws, r, 7, f[6]); put(ws, r, 8, f[8] if len(f) > 8 else f[7])
    r += 1
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:H{r-1}"

# =============================================================================
# Sheet 3 : Test cases / 测试用例
# =============================================================================
ws = wb.create_sheet("2-测试用例TestCase")
cols = ["TC-ID","用例名 Test Name","类型","覆盖特性","激励描述 Stimulus / Scenario",
        "检查点 Check","优先级","依赖","状态"]
set_widths(ws, [14, 26, 8, 16, 50, 30, 8, 12, 8])
ws.merge_cells("A1:I1")
put(ws, 1, 1, "测试用例清单 (Test Case List)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(cols, 1): put(ws, 2, i, h)
style_header(ws, 2, len(cols))

tc = [
    ("TC-SANITY","sanity_test","定向","F-FC-01/02","单笔 RD + 单笔 WR, 各字段固定值, 观察 AR/AW/W/R/B 与 rsp","端到端字段一致","P0","-","计划"),
    ("TC-RD-01","rd_single_test","定向","F-REQ-01,F-RSP-01","单笔 RD, Len 覆盖 1/8/64/256B, 对齐地址","araddr/arlen/rsp data 匹配","P0","SANITY","计划"),
    ("TC-RD-02","rd_narrow_unalign_test","定向","F-REQ-01","非对齐地址 + 窄传输 RD, 检查 wstrb/be","be/byte offset 正确","P0","RD-01","计划"),
    ("TC-WR-01","wr_single_test","定向","F-REQ-02,F-RSP-02","单笔 WR, Len 覆盖, wstrb 全/部分有效","wdata/wstrb/bresp 匹配","P0","SANITY","计划"),
    ("TC-BURST-01","burst_type_test","定向","F-REQ-04","INCR(1..256) 与 WRAP(2/4/8/16) 全组合","burst/size/len 正确","P0","RD-01","计划"),
    ("TC-ADDR-01","addr_map_test","随机","F-REQ-03","随机 subrange+addr, 覆盖 40bit 地址空间","映射地址正确","P0","-","计划"),
    ("TC-QOS-01","qos_test","随机","F-QOS-01","随机 QoS0..7 (bar-graph urgency)","awqos/arqos 匹配","P1","-","计划"),
    ("TC-OUT-01","outstanding_test","随机","F-OUT-01","连续下发至 outstanding 上限=8, 观察反压","ready 反压正确, 无丢包","P0","RD-01","计划"),
    ("TC-OOO-01","ooo_rsp_test","定向","F-OOO-01","AXI slave 乱序返回多笔响应","{AXID,Opc,TAG}索引匹配","P0","OUT-01","计划"),
    ("TC-ILV-01","interleave_rd_test","定向","F-ILV-01/02/03","双 ID 读交织返回(NBYTE=8), 按片段拆包","CONT/OK/LW/Addr 正确","P0","OOO-01","计划"),
    ("TC-AXID-01","axid_map_test","随机","F-AXID-01/02","OrderKey(8b) 覆盖>16 个值映射到 AXID(4b), 回填检查","AXID 映射+回填正确","P0","OUT-01","计划"),
    ("TC-AXID-BP","axid_exhaust_test","定向","F-AXID-03","占满全部 AXID/TAG, 验证上游反压与恢复","反压后无死锁","P1","AXID-01","计划"),
    ("TC-ERR-01","req_err_test","定向","F-ERR-01/02/04","注入 Status=ERR 请求(各 ErrorCode), 不下发 AXI","自组织 err rsp, code 一致","P0","SANITY","计划"),
    ("TC-ERR-MIX","err_okey_mix_test","随机","F-ERR-03","同 AXID 交替 err/okey 请求, 保序回响应","顺序与 fir/follo_err 正确","P0","ERR-01","计划"),
    ("TC-WRAP-01","wrap_align_test","定向","F-WRAP-01/02/03","非对齐 WRAP RD/WR (Len=2^n-1), 覆盖 RWRAP_CNT_MAX","对齐/修正/拼接正确","P0","BURST-01","计划"),
    ("TC-BP-WAR","bp_war_test","定向","F-BP-01","同 64B 块 写→读 序列","读等待写完成","P0","OUT-01","计划"),
    ("TC-BP-WAW","bp_waw_test","定向","F-BP-02","同块 写→写 序列","写序保持","P0","BP-WAR","计划"),
    ("TC-BP-RAW","bp_raw_test","定向","F-BP-03","同块 读→写 序列(BP=2)","写等待读完成","P0","BP-WAR","计划"),
    ("TC-BP-BLK","bp_block_gran_test","随机","F-BP-04","随机地址落在块边界/跨块","仅同块保序, 异块并发","P1","BP-WAR","计划"),
    ("TC-ELY-01","ely_buf_wr_test","定向","F-ELY-01/02","bufferable 写: 提前回 rsp + real rsp 旁路","early rsp 早于 real, 旁路","P0","WR-01","计划"),
    ("TC-ELY-NB","ely_nonbuf_test","定向","F-ELY-03","non-bufferable 写: 常规响应","无 early, 走 real rsp","P0","ELY-01","计划"),
    ("TC-ELY-WAR","ely_war_test","定向","F-ELY-04","early 写后读同地址","读等待写 real 完成","P1","ELY-01,BP-WAR","计划"),
    ("TC-WD-01","watchdog_timeout_test","定向","F-WD-01/03","slave 不回响应触发超时, 迟到 real rsp 丢弃","timeout err(110)+丢弃","P0","OUT-01","计划"),
    ("TC-WD-CNT","watchdog_cnt_test","定向","F-WD-02","验证 timer 双倍计数换算边界","计数换算正确","P1","WD-01","计划"),
    ("TC-RST-01","reset_test","定向","F-RST-01","仿真中随机 aresetn, 复位后重新激励","复位安全+可恢复","P0","-","计划"),
    ("TC-MIX-01","rd_wr_mix_random","随机","F-RSP-03,many","读写混合随机大流量回归","全 SB 检查通过","P0","many","计划"),
    ("TC-STRESS-01","backpressure_stress","随机","F-BP2-01","全通道 valid/ready 随机抖动 + 满 outstanding","无死锁/无丢包","P1","MIX-01","计划"),
    ("TC-STRESS-02","full_random_regress","随机","ALL","全特性混合约束随机, 长时回归收敛覆盖率","覆盖率达标","P0","ALL","计划"),
]
r = 3
for t in tc:
    put(ws, r, 1, t[0], bold_font)
    for i in range(1, 9):
        put(ws, r, i+1, t[i], bold_font if i in (6,) else cell_font,
            P0_FILL if (i == 6 and t[6] == "P0") else None,
            WRAP_C if i in (2,6,8) else WRAP_L)
    r += 1
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:I{r-1}"

# =============================================================================
# Sheet 4 : Env architecture / 验证环境架构
# =============================================================================
ws = wb.create_sheet("3-验证环境架构Env")
set_widths(ws, [26, 22, 74])
ws.merge_cells("A1:C1")
put(ws, 1, 1, "UVM 验证环境组件说明 (Environment Architecture)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(["组件 Component","层级/类型","职责 Responsibility"], 1): put(ws, 2, i, h)
style_header(ws, 2, 3)
env = [
    ("tb_top","module","实例化 DUT(axi_tniu)、rknp_if、axi_if, 生成时钟复位, run_test()"),
    ("rknp_if","interface","RKNP 流控层接口(head/tail/valid/ready/data), HeadPenalty=0, 含协议断言"),
    ("axi_if","interface","AXI4 接口(AW/W/AR/R/B 五通道), 含 AXI 协议断言"),
    ("rknp_agent","uvm_agent(active,master)","向 DUT 驱动 req packet, 采样 req/rsp packet"),
    ("  rknp_driver","uvm_driver","按 HeadPenalty=0 时序打包并驱动 req flit; 采集 rsp ready 反压"),
    ("  rknp_monitor","uvm_monitor","采样 rxreq(req) 与 txrsp(rsp) 两条流, 解包为事务发往 SB/cov"),
    ("  rknp_sequencer","uvm_sequencer","调度 rknp_seq_item"),
    ("axi_agent","uvm_agent(active,slave)","作为 AXI slave 响应 DUT master, 支持乱序/交织/outstanding/延迟注入"),
    ("  axi_driver","uvm_driver(slave)","响应 AR/AW/W, 生成 R/B; 可配 out-of-order、interleave、随机延迟"),
    ("  axi_monitor","uvm_monitor","采样 AW/W/AR/R/B 事务发往 SB/cov"),
    ("  axi_sequencer","uvm_sequencer","调度 slave 响应策略(resp/delay/reorder)"),
    ("axi_tniu_env","uvm_env","组织 agent、scoreboard、refmodel、coverage、virtual sequencer"),
    ("  virtual_sequencer","uvm_sequencer","跨 agent 协调(v-seq 同时控制 rknp 激励与 axi slave 策略)"),
    ("axi_tniu_refmodel","uvm_component","参考模型/predictor: 由 req 预测应产生的 AXI 事务与最终 rsp packet"),
    ("axi_tniu_scoreboard","uvm_scoreboard","req->AXI 转换检查 + AXI rsp->rsp packet 检查 + 保序/交织/error/timeout 检查"),
    ("axi_tniu_coverage","uvm_subscriber","功能覆盖 covergroup 采样(见 5-覆盖率计划)"),
    ("axi_tniu_cfg","uvm_object","环境配置: 使能开关(err/wrap/early/bp/axidmap)、位宽参数、延迟/乱序策略"),
    ("rknp_pkg","package","RKNP req/rsp flit 字段偏移/位宽定义 + pack/unpack 函数(单一数据契约来源)"),
    ("seq_lib","package","各特性对应 sequence(定向) + 约束随机 sequence + virtual sequence"),
    ("test_lib","package","base_test + 各 directed/random test, 通过 +UVM_TESTNAME 选择"),
]
r = 3
for c in env:
    put(ws, r, 1, c[0], bold_font if not c[0].startswith("  ") else cell_font)
    put(ws, r, 2, c[1], cell_font, None, WRAP_C); put(ws, r, 3, c[2])
    r += 1

# =============================================================================
# Sheet 5 : Assertions & checks / 断言检查点
# =============================================================================
ws = wb.create_sheet("4-断言检查点Checks")
set_widths(ws, [14, 20, 66, 12])
ws.merge_cells("A1:D1")
put(ws, 1, 1, "协议断言与 Scoreboard 检查点 (Assertions & Checks)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(["Check-ID","位置 Location","检查内容 Description","类型"], 1): put(ws, 2, i, h)
style_header(ws, 2, 4)
chk = [
    ("A-RKNP-01","rknp_if(SVA)","valid 拉高后到 ready 前 data/head/tail 保持稳定","assertion"),
    ("A-RKNP-02","rknp_if(SVA)","head 仅在每个 packet 首拍有效; tail 标识最后一拍 body","assertion"),
    ("A-RKNP-03","rknp_if(SVA)","HeadPenalty=0: head 与首拍 body 同拍出现","assertion"),
    ("A-RKNP-04","rknp_if(SVA)","复位期间 valid 不得为 1","assertion"),
    ("A-AXI-01","axi_if(SVA)","AXVALID 拉高后至 AXREADY 前地址通道信号保持稳定","assertion"),
    ("A-AXI-02","axi_if(SVA)","WLAST 与 awlen 拍数一致; 每 burst 恰一个 WLAST","assertion"),
    ("A-AXI-03","axi_if(SVA)","RLAST 与 arlen 拍数一致","assertion"),
    ("A-AXI-04","axi_if(SVA)","B/R 的 ID 必须是曾发出的 outstanding ID","assertion"),
    ("A-AXI-05","axi_if(SVA)","arsize/awsize == $clog2(NBYTEPERWORD)=3","assertion"),
    ("C-CONV-01","scoreboard","每个 RD req -> 恰一组 AR 事务, 地址/len/burst/size/id(映射后) 一致","scoreboard"),
    ("C-CONV-02","scoreboard","每个 WR req -> 恰一组 AW+W, wdata/wstrb 与 body 一致","scoreboard"),
    ("C-CONV-03","scoreboard","每组 R -> read rsp packet, data/be/LW/addr 一致","scoreboard"),
    ("C-CONV-04","scoreboard","每个 B -> write rsp packet, status/errcode 一致","scoreboard"),
    ("C-AXID-01","scoreboard","OrderKey->AXID 映射一致; rsp 回填 OrderKey == 原始请求","scoreboard"),
    ("C-ORD-01","scoreboard","同 AXID+Opc 的响应顺序与请求下发顺序一致(TAG_CNT)","scoreboard"),
    ("C-ORD-02","scoreboard","乱序返回场景下 rsp packet 与请求正确配对","scoreboard"),
    ("C-ILV-01","scoreboard","读交织: 拆包后各 packet Status(CONT/OK)/Addr/LW 正确","scoreboard"),
    ("C-BP-01","scoreboard","同地址块保序: 依赖被违反则报错(WAR/WAW/RAW 按 BP_TYPE)","scoreboard"),
    ("C-ERR-01","scoreboard","Status=ERR 请求不产生任何 AXI 事务","scoreboard"),
    ("C-ERR-02","scoreboard","error rsp 的 ErrorCode/Status 与请求一致","scoreboard"),
    ("C-ELY-01","scoreboard","bufferable 写 early rsp 时序早于其 real B; real B 被旁路不二次回包","scoreboard"),
    ("C-WD-01","scoreboard","超时请求返回 ErrorCode=110(Time-out) 的 error rsp","scoreboard"),
    ("C-WD-02","scoreboard","超时后到达的 real response 被丢弃, 不产生重复 rsp","scoreboard"),
    ("C-WRAP-01","scoreboard","非对齐 WRAP: 下发 AXI 地址对齐, 回程数据修正回原始顺序","scoreboard"),
    ("C-LEAK-01","scoreboard","仿真结束无未匹配的挂起事务(req/rsp/AXI 全部配平)","scoreboard"),
]
r = 3
for c in chk:
    put(ws, r, 1, c[0], bold_font); put(ws, r, 2, c[1]); put(ws, r, 3, c[2])
    put(ws, r, 4, c[3], cell_font, None, WRAP_C); r += 1
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:D{r-1}"

# =============================================================================
# Sheet 6 : Coverage plan / 覆盖率计划
# =============================================================================
ws = wb.create_sheet("5-覆盖率计划Coverage")
set_widths(ws, [22, 26, 58, 12])
ws.merge_cells("A1:D1")
put(ws, 1, 1, "功能覆盖率计划 (Functional Coverage Plan)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(["Covergroup","Coverpoint / Cross","Bins / 说明","目标"], 1): put(ws, 2, i, h)
style_header(ws, 2, 4)
cov = [
    ("cg_req","cp_opc","RD, RDW, WR, WRW 四类","100%"),
    ("cg_req","cp_len","Len bins: 1,2,4,8,15,16,63,64,127,128,255,256B","100%"),
    ("cg_req","cp_addr_align","aligned / unaligned(within 64bit word)","100%"),
    ("cg_req","cp_qos","QoS 0..7","100%"),
    ("cg_req","cp_status","OK / ERR","100%"),
    ("cg_req","cp_errcode","ERR 时 ErrorCode 000..110","100%"),
    ("cg_req","cross opc x len x align","读写 x 长度 x 对齐","关键组合"),
    ("cg_axid","cp_orderkey","OrderKey 全 8bit 分区 (>=16 值)","100%"),
    ("cg_axid","cp_axid","映射后 AXID 0..15","100%"),
    ("cg_axid","cp_tag_cnt","TAG_CNT 0..SUP_REQ_NUM-1","100%"),
    ("cg_axid","cross orderkey x axid","覆盖多 OrderKey 复用同 AXID","命中"),
    ("cg_burst","cp_arburst/awburst","INCR / WRAP","100%"),
    ("cg_burst","cp_wrap_len","WRAP 2/4/8/16 beats","100%"),
    ("cg_burst","cp_size","awsize/arsize == 3","100%"),
    ("cg_ooo","cp_rsp_order","in-order / out-of-order 返回","100%"),
    ("cg_ooo","cp_outstanding","当前 outstanding 深度 1..8","100%"),
    ("cg_ilv","cp_interleave","无交织 / 2-ID 交织 / 多-ID 交织","100%"),
    ("cg_ilv","cp_cont_ok","rsp Status CONT / OK 序列","100%"),
    ("cg_err","cp_err_pos","单笔err / 连续err / err后okey / okey后err","100%"),
    ("cg_err","cp_err_type","各 ErrorCode 注入","100%"),
    ("cg_bp","cp_hazard","WAR / WAW / RAW / 无冒险","100%"),
    ("cg_bp","cp_block","同块 / 跨块 / 块边界","100%"),
    ("cg_ely","cp_bufferable","bufferable / non-bufferable 写","100%"),
    ("cg_ely","cp_ely_seq","early先/real先 时序关系","100%"),
    ("cg_ely","cross bufferable x war","early 写 + 后续同地址读","命中"),
    ("cg_wrap","cp_wrap_align","对齐 / 非对齐 WRAP","100%"),
    ("cg_wrap","cp_rwrap_cnt","拆分计数 1..RWRAP_CNT_MAX(4)","100%"),
    ("cg_wd","cp_timeout","正常 / 超时","100%"),
    ("cg_wd","cp_late_rsp","超时后 real rsp 到达 / 未到达","100%"),
    ("cg_resp","cp_bresp/rresp","OKAY/EXOKAY/SLVERR/DECERR","100%"),
]
r = 3
prev = None
for c in cov:
    fill = SUB_FILL if c[0] != prev else None
    put(ws, r, 1, c[0], bold_font, fill); put(ws, r, 2, c[1]); put(ws, r, 3, c[2])
    put(ws, r, 4, c[3], cell_font, None, WRAP_C)
    prev = c[0]; r += 1
ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:D{r-1}"

# =============================================================================
# Sheet 7 : Checklist / 验证进度清单
# =============================================================================
ws = wb.create_sheet("6-验证进度清单Checklist")
set_widths(ws, [10, 46, 14, 14, 26])
ws.merge_cells("A1:E1")
put(ws, 1, 1, "验证里程碑与进度清单 (Milestones & Checklist)", title_font, H_FILL, WRAP_C)
ws.row_dimensions[1].height = 26
for i, h in enumerate(["阶段","任务 Task","负责人 Owner","状态 Status","备注 Note"], 1): put(ws, 2, i, h)
style_header(ws, 2, 5)
ck = [
    ("M1-规划","验证计划评审通过","","计划",""),
    ("M1-规划","接口/字段契约(rknp_pkg)与设计对齐确认","","计划","关键:与子模块提取一致"),
    ("M2-环境","搭建 tb_top + 接口 + 时钟复位","","计划",""),
    ("M2-环境","RKNP master agent 完成","","计划",""),
    ("M2-环境","AXI slave agent 完成(含乱序/交织/延迟)","","计划",""),
    ("M2-环境","scoreboard + refmodel 打通 sanity","","计划",""),
    ("M2-环境","接口协议断言接入","","计划",""),
    ("M3-基础","读/写/突发/地址映射 定向用例通过","","计划","P0"),
    ("M3-基础","outstanding / 乱序 / 交织 通过","","计划","P0"),
    ("M4-可选","AXID 映射 用例通过","","计划","P0"),
    ("M4-可选","Req-Error 处理 用例通过","","计划","P0"),
    ("M4-可选","Wrap 重排 用例通过","","计划","P0"),
    ("M4-可选","同地址保序(WAR/WAW/RAW) 用例通过","","计划","P0"),
    ("M4-可选","Early/Bufferable 用例通过","","计划","P0"),
    ("M4-可选","Watchdog 超时 用例通过","","计划","P0"),
    ("M5-回归","随机回归 0 error/0 fatal","","计划",""),
    ("M5-回归","功能覆盖率 >= 100%","","计划",""),
    ("M5-回归","代码覆盖率 >= 98% + 排除评审","","计划",""),
    ("M5-回归","验证报告与签核","","计划",""),
]
r = 3
for c in ck:
    put(ws, r, 1, c[0], bold_font, SUB_FILL, WRAP_C)
    put(ws, r, 2, c[1]); put(ws, r, 3, c[2], cell_font, None, WRAP_C)
    put(ws, r, 4, c[3], cell_font, CFG_FILL, WRAP_C); put(ws, r, 5, c[4])
    r += 1
ws.freeze_panes = "A3"

import os
os.makedirs("/home/claude/axi_tniu_uvm/doc", exist_ok=True)
out = "/home/claude/axi_tniu_uvm/doc/axi_tniu_vplan.xlsx"
wb.save(out)
print("saved", out)
